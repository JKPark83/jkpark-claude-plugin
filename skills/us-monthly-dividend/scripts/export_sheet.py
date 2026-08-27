#!/usr/bin/env python3
"""Build a styled .xlsx portfolio report for upload to Google Drive.

Usage: python3 export_sheet.py report.json out.xlsx

Styling follows the Anthropic xlsx-skill number-format conventions
($#,##0.00, percents, negative handling) plus common openpyxl report
patterns (colored section/table headers, zebra rows, thin borders,
bold totals). Formatting survives the Drive xlsx -> Google Sheets
conversion. Section markers ([요약], [보유종목], ...) and the holdings
header are kept identical to the CSV era so a later rebalance run can
still parse the sheet back as account state.

Input JSON schema (null = unknown, cell left blank):
{
  "title": str,
  "summary": [{"label": str, "value": str|num, "fmt":
               "text"|"usd"|"krw"|"pct"|"fx"|"int"}, ...],
  "holdings": [{"name","ticker","target_pct","band_lo","band_hi",
                "avg_cost","shares","cost_usd","price","value_usd",
                "gain_usd","gain_pct","weight_pct","ttm_yield",
                "rebalance"}, ...],
  "totals_usd": {"cost_usd","value_usd","weight_pct","ttm_yield"},
  "totals_krw": {"value_krw"},
  "calendar": [{"name", "months": [12 nums], "annual"}, ...],
  "calendar_total": {"months": [12 nums], "annual"},
  "cashflow": {"usd": [12], "usd_avg", "krw": [12], "krw_avg"},
  "commentary": [{"heading": str, "body": str}, ...],   # optional
  "footnote": str
}

commentary is free-form prose: one entry per topic (시장 상황 평가,
종목 선정 이유, 백테스트 해석, ...). body may contain newlines; each
entry renders as a bold heading row plus a merged wrapped-text block.
"""
import json
import sys

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.stderr.write("openpyxl not installed. Run: python3 -m pip install openpyxl\n")
    sys.exit(2)

# palette
NAVY_DARK = "1F4E79"   # title
NAVY = "4472C4"        # summary/holdings section + header
NAVY_LIGHT = "DDEBF7"  # totals fill
GREEN = "548235"       # calendar section + header
GREEN_LIGHT = "E2EFDA"
ORANGE = "C55A11"      # cashflow section + header
ORANGE_LIGHT = "FCE4D6"
PURPLE = "7030A0"      # commentary section
PURPLE_LIGHT = "E4DFEC"
ZEBRA = "F2F2F2"
GAIN_RED = "C00000"    # Korean convention: gain red, loss blue
LOSS_BLUE = "0070C0"
GRAY = "808080"

FMT = {
    "usd": '$#,##0.00',
    "krw": '₩#,##0',
    "pct": '0.00"%"',
    "fx": '#,##0.00',
    "int": '#,##0',
}
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NCOLS = 15  # widest table (holdings)


def fill(color):
    return PatternFill("solid", fgColor=color)


class Sheet:
    def __init__(self, ws):
        self.ws = ws
        self.row = 0

    def next(self):
        self.row += 1
        return self.row

    def put(self, r, c, value, fmt=None, font=None, fg=None, border=True,
            align=None, bold=False, color=None):
        cell = self.ws.cell(row=r, column=c)
        if value is not None:
            cell.value = value
        if fmt:
            cell.number_format = FMT.get(fmt, fmt)
        cell.font = font or Font(name="Arial", size=10, bold=bold,
                                 color=color or "000000")
        if fg:
            cell.fill = fill(fg)
        if border:
            cell.border = BOX
        if align:
            cell.alignment = Alignment(horizontal=align, vertical="center")
        return cell

    def title_row(self, text):
        r = self.next()
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r,
                            end_column=NCOLS)
        self.put(r, 1, text, fg=NAVY_DARK, border=False, align="center",
                 font=Font(name="Arial", size=14, bold=True, color="FFFFFF"))
        self.ws.row_dimensions[r].height = 26

    def section_row(self, text, color):
        r = self.next()
        self.ws.merge_cells(start_row=r, start_column=1, end_row=r,
                            end_column=NCOLS)
        self.put(r, 1, text, fg=color, border=False,
                 font=Font(name="Arial", size=11, bold=True, color="FFFFFF"))
        self.ws.row_dimensions[r].height = 20

    def header_row(self, labels, color):
        r = self.next()
        for c, label in enumerate(labels, start=1):
            self.put(r, c, label, fg=color, align="center",
                     font=Font(name="Arial", size=10, bold=True,
                               color="FFFFFF"))
            self.ws.cell(row=r, column=c).alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
        self.ws.row_dimensions[r].height = 28
        return r


def signed_font(value, bold=False):
    if value is None:
        return Font(name="Arial", size=10, bold=bold)
    color = GAIN_RED if value > 0 else (LOSS_BLUE if value < 0 else "000000")
    return Font(name="Arial", size=10, bold=bold, color=color)


def main():
    if len(sys.argv) != 3:
        sys.stderr.write("usage: export_sheet.py report.json out.xlsx\n")
        sys.exit(1)
    with open(sys.argv[1], encoding="utf-8") as f:
        data = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "포트폴리오"
    ws.sheet_view.showGridLines = False
    s = Sheet(ws)

    widths = [26, 9] + [11] * (NCOLS - 2)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    s.title_row(data["title"])
    s.next()  # spacer

    # [요약]
    s.section_row("[요약]", NAVY)
    for i, item in enumerate(data["summary"]):
        r = s.next()
        zebra = ZEBRA if i % 2 else None
        s.put(r, 1, item["label"], fg=zebra, bold=True)
        fmt = item.get("fmt", "text")
        s.put(r, 2, item["value"], fmt=None if fmt == "text" else fmt,
              fg=zebra, align="left" if fmt == "text" else "right")
    s.next()

    # [보유종목]
    s.section_row("[보유종목]", NAVY)
    s.header_row(["종목", "Ticker", "목표비중(%)", "밴드하한(%)", "밴드상한(%)",
                  "평단(USD)", "주수", "매입금액(USD)", "현재단가(USD)",
                  "평가금액(USD)", "수익금(USD)", "수익률(%)", "현재비중(%)",
                  "TTM배당률(%)", "리밸런싱주수"], NAVY)
    for i, h in enumerate(data["holdings"]):
        r = s.next()
        zebra = ZEBRA if i % 2 else None
        s.put(r, 1, h["name"], fg=zebra)
        s.put(r, 2, h["ticker"], fg=zebra, align="center", bold=True)
        for c, key, fmt in ((3, "target_pct", "pct"), (4, "band_lo", "pct"),
                            (5, "band_hi", "pct"), (6, "avg_cost", "usd"),
                            (7, "shares", "int"), (8, "cost_usd", "usd"),
                            (9, "price", "usd"), (10, "value_usd", "usd"),
                            (13, "weight_pct", "pct"),
                            (14, "ttm_yield", "pct")):
            s.put(r, c, h.get(key), fmt=fmt, fg=zebra, align="right")
        for c, key, fmt in ((11, "gain_usd", "usd"), (12, "gain_pct", "pct")):
            v = h.get(key)
            s.put(r, c, v, fmt=fmt, fg=zebra, align="right",
                  font=signed_font(v))
        reb = h.get("rebalance")
        reb_num = None
        if isinstance(reb, str) and reb.lstrip("+-").isdigit():
            reb_num = int(reb)
        s.put(r, 15, reb, fg=zebra, align="center",
              font=signed_font(reb_num, bold=True))
    tu = data["totals_usd"]
    r = s.next()
    s.put(r, 1, "합계(USD)", fg=NAVY_LIGHT, bold=True)
    for c in range(2, NCOLS + 1):
        s.put(r, c, None, fg=NAVY_LIGHT)
    s.put(r, 8, tu.get("cost_usd"), fmt="usd", fg=NAVY_LIGHT, bold=True,
          align="right")
    s.put(r, 10, tu.get("value_usd"), fmt="usd", fg=NAVY_LIGHT, bold=True,
          align="right")
    s.put(r, 13, tu.get("weight_pct"), fmt="pct", fg=NAVY_LIGHT, bold=True,
          align="right")
    s.put(r, 14, tu.get("ttm_yield"), fmt="pct", fg=NAVY_LIGHT, bold=True,
          align="right")
    r = s.next()
    s.put(r, 1, "합계(KRW)", fg=NAVY_LIGHT, bold=True)
    for c in range(2, NCOLS + 1):
        s.put(r, c, None, fg=NAVY_LIGHT)
    s.put(r, 10, data["totals_krw"].get("value_krw"), fmt="krw",
          fg=NAVY_LIGHT, bold=True, align="right")
    s.next()

    # [월별 배당 달력(세전 USD)]
    s.section_row("[월별 배당 달력(세전 USD)]", GREEN)
    s.header_row(["종목"] + [f"{m}월" for m in range(1, 13)] + ["연간"], GREEN)
    for i, row in enumerate(data["calendar"]):
        r = s.next()
        zebra = ZEBRA if i % 2 else None
        s.put(r, 1, row["name"], fg=zebra)
        for c, v in enumerate(row["months"], start=2):
            s.put(r, c, v, fmt="usd", fg=zebra, align="right")
        s.put(r, 14, row["annual"], fmt="usd", fg=zebra, bold=True,
              align="right")
    ct = data["calendar_total"]
    r = s.next()
    s.put(r, 1, "합계", fg=GREEN_LIGHT, bold=True)
    for c, v in enumerate(ct["months"], start=2):
        s.put(r, c, v, fmt="usd", fg=GREEN_LIGHT, bold=True, align="right")
    s.put(r, 14, ct["annual"], fmt="usd", fg=GREEN_LIGHT, bold=True,
          align="right")
    s.next()

    # [세후 월 현금흐름]
    s.section_row("[세후 월 현금흐름]", ORANGE)
    s.header_row(["통화"] + [f"{m}월" for m in range(1, 13)] + ["월평균"],
                 ORANGE)
    cf = data["cashflow"]
    for i, (label, key, avg_key, fmt) in enumerate(
            (("세후 USD", "usd", "usd_avg", "usd"),
             ("세후 KRW", "krw", "krw_avg", "krw"))):
        r = s.next()
        zebra = ZEBRA if i % 2 else None
        s.put(r, 1, label, fg=zebra, bold=True)
        for c, v in enumerate(cf[key], start=2):
            s.put(r, c, v, fmt=fmt, fg=zebra, align="right")
        s.put(r, 14, cf[avg_key], fmt=fmt, fg=ORANGE_LIGHT, bold=True,
              align="right")

    # [의견]
    if data.get("commentary"):
        s.next()
        s.section_row("[의견]", PURPLE)
        for item in data["commentary"]:
            r = s.next()
            ws.merge_cells(start_row=r, start_column=1, end_row=r,
                           end_column=NCOLS)
            s.put(r, 1, item["heading"], fg=PURPLE_LIGHT, border=False,
                  font=Font(name="Arial", size=10, bold=True))
            body = item.get("body", "")
            r = s.next()
            ws.merge_cells(start_row=r, start_column=1, end_row=r,
                           end_column=NCOLS)
            cell = s.put(r, 1, body, border=False,
                         font=Font(name="Arial", size=10))
            cell.alignment = Alignment(horizontal="left", vertical="top",
                                       wrap_text=True)
            # merged wrapped cells don't auto-size: estimate the height
            # (~80 chars per line across the merged width, Korean-safe)
            lines = sum(max(1, -(-len(p) // 80))
                        for p in body.split("\n")) or 1
            ws.row_dimensions[r].height = lines * 15 + 6

    if data.get("footnote"):
        s.next()
        r = s.next()
        ws.merge_cells(start_row=r, start_column=1, end_row=r,
                       end_column=NCOLS)
        s.put(r, 1, data["footnote"], border=False,
              font=Font(name="Arial", size=8, italic=True, color=GRAY))

    wb.save(sys.argv[2])
    print(f"saved {sys.argv[2]}")


if __name__ == "__main__":
    main()
